import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def exportar_simulacion_a_excel(vectores_por_dia, ruta="simulacion_peluqueria.xlsx"):
    escritor = pd.ExcelWriter(ruta, engine='xlsxwriter')

    for i, vector in enumerate(vectores_por_dia):
        filas = []
        for fila in vector:
            fila_salida = dict(fila)
            clientes = fila_salida.pop("clientes_activos", [])
            for cliente in clientes:
                col = f"cliente_{cliente['id']}"
                desc = f"{cliente['estado']} ({cliente['hora_refrigerio']})" if cliente['hora_refrigerio'] else cliente['estado']
                fila_salida[col] = desc
            filas.append(fila_salida)

        df = pd.DataFrame(filas)

        # Redondear flotantes
        for col in df.select_dtypes(include=['float', 'float64']).columns:
            df[col] = df[col].round(2)

        sheet_name = f"Día {i+1}"
        df.to_excel(escritor, sheet_name=sheet_name, index=False)

    escritor.close()

    # Ajuste de columnas usando openpyxl
    wb = load_workbook(ruta)
    for ws in wb.worksheets:
        for col_idx, col in enumerate(ws.iter_cols(1, ws.max_column), 1):
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            adjusted_width = max(12, min(max_len * 1.4, 80))  # puedes ajustar el 1.4 si quieres más espacio
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    wb.save(ruta)